from tkinter import *
from math import *
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
from openpyxl import load_workbook

LEN = 6
NO = 0
CODE = 1
NAME = 2
PRICE = 3
QUAN = 4
TOTAL = 5

df = 0
item_list = 0
current_pointer = 0
save_barcode = 0

master = Tk()
master.geometry("1500x1000")
master.title("TAP HOA THU&HUONG")

fields = 'MA VACH', 'TEN HANG HOA', 'GIA CA'

#use arrow  keys to move around the quantity entry boxes
def arrow_key_up(labels):
   print('arrow_key_up() called')
   global current_pointer
   if current_pointer > 0:
      current_pointer -= 1

   (labels[current_pointer][QUAN]).focus()

#use arrow  keys to move around the quantity entry boxes
def arrow_key_down(labels):
   print('arrow_key_down() called')
   global current_pointer
   global item_list

   if current_pointer < item_list-1:
      current_pointer += 1
   (labels[current_pointer][QUAN]).focus()
#write info to screen
def write(strr):
   T.delete('1.0', END)
   T.insert(END, "\n"+strr, 'big')
#quit the program when press escape   
def onclickQuit(event=None):
    print("You clicked the QUIT button")
    master.destroy()
def barcode_focus(ents):
   ents['MA VACH'].delete(0,END)
   ents['MA VACH'].focus()
#make form for detail - upper part
def makeform(master, fields):
   entries = {}
   for field in fields:
      row = Frame(master)
      lab = Label(row, width=15, text=field+": ", anchor='w', font = "Verdana 20 bold")
      ent = Entry(row,  font = "Verdana 30")
      ent.insert(0,"0")
      row.pack(side=TOP, fill=X, padx=5, pady=5)
      lab.pack(side=LEFT)
      ent.pack(side=RIGHT, expand=YES, fill=X)
      entries[field] = ent
      if field == 'MA VACH':#add the label RECENT_CODE to gloabl entries to show the last very recent barcode
         entries['RECENT_CODE'] = lab   
   return entries

def init():
   print("init() called")
   global df
   df = pd.read_excel('code.xlsx', sheet_name='HangHoa')

def reset(entries, labels):
   #delete the detail price
   print("You clicked the RESET button")
   write("Bắt đầu tính tiền khách mới \nwelcome to the Earth\n **hugging**")
   entries['MA VACH'].delete(0,END)
   entries['MA VACH'].insert(0, '0')
   entries['TEN HANG HOA'].delete(0,END)
   entries['TEN HANG HOA'].insert(0, '0')
   entries['GIA CA'].delete(0,END)
   entries['GIA CA'].insert(0, '0')


   t1.config(text='TỔNG CỘNG: 0')

   #reset the list of purchase items
   global item_list 
   for r in range(0,item_list):
      for c in range(0,LEN):
         (labels[r][c]).config(text=".")
         if c == QUAN:
            (labels[r][c]).delete(0,END)
            (labels[r][c]).insert(0, "1")

   item_list = 0
   current_pointer = 0
   save_barcode = 0

   barcode_focus(entries)


def enter_pressed(entries, labels):
   global save_barcode
   print('enter_pressed() called')
   if entries['MA VACH'].get() == "":#for update quantity
      input_quantity(labels)
   else:#for new barcode
      save_barcode = entries['MA VACH'].get()
      entries['RECENT_CODE'].config(text="MA VACH: \n ("+str(save_barcode) +")")
      sum(entries)

   barcode_focus(entries)   

#update quantity at specific entry item and re-calculate the final_total_price
def input_quantity(labels):
   print ('input_quantity() called')
   global item_list 
   for r in range(0,item_list):
      final_price =  float((labels[r][PRICE]).cget("text")) * float((labels[r][QUAN]).get())
      (labels[r][TOTAL]).config(text=str(final_price))
   price_sum(labels)

def sum(entries):
   print('sum() called')
   global item_list
   global current_pointer
   flag = print_out_detail(entries)
   if flag:      
      print_out_item(entries)
   else:
      write("Chưa có mặt hàng này :( \nNot in database T.T")

      entries['TEN HANG HOA'].delete(0,END)
      entries['TEN HANG HOA'].insert(0, "Chưa có")
      entries['GIA CA'].delete(0,END)
      entries['GIA CA'].insert(0, "Chưa có")
#calculate total price   
def price_sum(labels):
   # write("Tính Tiền nè\nWe love you ^^ ")
   s = 0
   for r in range(0, item_list):      
      s += float((labels[r][TOTAL]).cget("text"))
   
   t1.config(text='TỔNG CỘNG: ' + str(s))


#after add new item, print it to items list
def print_out_item(entries):
   print("print_out_item() called")
   global item_list
   global current_pointer
   global labels

   if check_current_list(entries):
      return 

   item_list += 1
   current_pointer = item_list
   update_product_list(labels,entries)
   price_sum(labels)

def check_current_list(entries):
   global labels
   for r in range(0, item_list): 
      # if entries['TEN HANG HOA'].get() == labels[r][NAME].cget("text"):
      if save_barcode == labels[r][CODE].cget("text"):
         print ("found same product in list")
         write("Vừa cập nhật số lượng cho: \n" + entries['TEN HANG HOA'].get())
         s = int(labels[r][QUAN].get()) + 1
         #update quantity
         labels[r][QUAN].delete(0,END)
         labels[r][QUAN].insert(0, str(s))
         #update price*new_quantity for that item
         final_price = float(entries['GIA CA'].get()) * float(labels[r][QUAN].get())
         (labels[r][TOTAL]).config(text=str(final_price))
         #recalculate total price
         price_sum(labels)
         return True
      
   return False
#update new or amend old data in database
def database_modify(entries):
   print("database_modify() called")
   flag = False
   wb = load_workbook('code.xlsx')
   ws = wb['HangHoa']
   last_idx = df['MaVach'].size
   
   for i in range(1,last_idx):
      if df['MaVach'][i] == int(save_barcode):
         print ('updated price of product')
         write("Đã Cập Nhật Lại Mặt Hàng :)")
         flag = True
         ws['C'+str(i+2)] = entries['TEN HANG HOA'].get()
         ws['E'+str(i+2)] = entries['GIA CA'].get()
         wb.save('code.xlsx')
         prettify(entries)
         break

   #for input new data
   if flag == False:
      print ('added new data')
      write("Đã Thêm Danh Mục Hàng Mới ^^")
      ws['A'+str(last_idx+2)] = last_idx+1
      ws['B'+str(last_idx+2)] = save_barcode #ma vach
      ws['C'+str(last_idx+2)] = entries['TEN HANG HOA'].get()
      ws['E'+str(last_idx+2)] = entries['GIA CA'].get()
      wb.save('code.xlsx')
      print_out_item(entries)#print just added item to bought-item-list

   init()#read new excel file
   barcode_focus(entries)

def prettify(entries):
   global labels
   print ('prettify() called')
   #update price*new_quantity for that item
   for r in range(0, item_list): 
       if save_barcode == labels[r][CODE].cget("text"):
          #update name and price
         (labels[r][NAME]).config(text=entries['TEN HANG HOA'].get())
         (labels[r][PRICE]).config(text=entries['GIA CA'].get())
         #update price and quantity
         final_price = float(labels[r][PRICE].cget("text")) * float(labels[r][QUAN].get())
         (labels[r][TOTAL]).config(text=str(final_price))
         #recalculate total price
         price_sum(labels)
         break

#print out products detail below - the info of below part
def update_product_list(labels,entries):
   print ('update_product_list() called')
   write("Đã tính thêm sản phẩm: \n" + entries['TEN HANG HOA'].get())
   for r in range(item_list-1, item_list):
      (labels[r][NO]).config(text=str(r+1))
      (labels[r][CODE]).config(text=str(save_barcode))
      (labels[r][NAME]).config(text=entries['TEN HANG HOA'].get())
      (labels[r][PRICE]).config(text=entries['GIA CA'].get())
      
      final_price = float(entries['GIA CA'].get()) * float((labels[r][QUAN]).get())
      (labels[r][TOTAL]).config(text=str(final_price))

def make_list_title():
   row = Frame(master)

   b = Label(row, width=5, text="STT", anchor='w',bg = "black", fg = "white", font = "Verdana 20 ")#position 0: no 1: name
   b.grid(row=0, column=NO) # for cases 0, 1, 2   

   b = Label(row, width=10, text="Mã Vạch", anchor='w',bg = "black", fg = "white", font = "Verdana 20")#position 0: no 1: name
   b.grid(row=0, column=CODE) # for cases 0, 1, 2

   b = Label(row, width=20, text="Tên Sản Phẩm", anchor='w',bg = "black", fg = "white", font = "Verdana 20")#position 0: no 1: name
   b.grid(row=0, column=NAME) # for cases 0, 1, 2

   b = Label(row, width=10, text="Giá", anchor='e',bg = "black", fg = "white", font = "Verdana 20") #position 2: price, j==3 is total price
   b.grid(row=0, column=PRICE) # for cases 0, 1, 2

   b = Label(row, width=10, text="Số Lượng", anchor='e',bg = "black", fg = "white",font = "Verdana 20") #position 2: price, j==3 is total price
   b.grid(row=0, column=QUAN) # for cases 0, 1, 2

   b = Label(row, width=10, text="Thành Tiền", anchor='e',bg = "black", fg = "white", font = "Verdana 20") #position 2: price, j==3 is total price
   b.grid(row=0, column=TOTAL) # for cases 0, 1, 2
   row.pack()


def init_list_detail():
   print('init_list_detail() called')
   T = [["","" , "", "", "", ""],["", "", "", "", "", ""],["",  "", "", "", "", ""],[ "","", "", "", "", ""],[ "",   "", "", "", "", ""],[ "","", "", "", "", ""], [ "","", "", "", "", ""],[ "","", "", "", "", ""],[ "","", "", "", "", ""],[ "","", "", "", "", ""]]
   make_list_title()

   height = 10
   for i in range(0,height): #Rows
      row = Frame(master)
      b = Label(row, width=5, text=".", anchor='w', font = "Verdana 20")#position 0: no 1: name
      b.grid(row=i, column=NO) # for cases 0, 1, 2
      T[i][NO] = b

      b = Label(row, width=10, text=".", anchor='w', font = "Verdana 20")#position 0: no 1: name
      b.grid(row=i, column=CODE) # for cases 0, 1, 2
      T[i][CODE] = b

      b = Label(row, width=20, text=".", anchor='w', font = "Verdana 20 bold")#position 0: no 1: name
      b.grid(row=i, column=NAME) # for cases 0, 1, 2
      T[i][NAME] = b

      b = Label(row, width=10, text=".", anchor='e', font = "Verdana 20 bold") #position 2: price, j==3 is total price
      b.grid(row=i, column=PRICE) # for cases 0, 1, 2
      T[i][PRICE] = b

      c = Entry(row, width=5, font = "Verdana 20")
      c.insert(0,"1")
      c.grid(row=i, column=QUAN)#entry is in position 3
      T[i][QUAN] = c #entry is in position 3

      b = Label(row, width=10, text=".", anchor='e', font = "Verdana 20") #position 2: price, j==3 is total price
      b.grid(row=i, column=TOTAL) # for cases 0, 1, 2
      T[i][TOTAL] = b

      row.pack()

   return T

def print_out_detail(entries):
   print ('print_out_detail() called')

   for i in range(0,df['MaVach'].size):#TESTING
      if df['MaVach'][i] == int(entries['MA VACH'].get()):
         
         print('FOUND IT')
         print (df['TenHangHoa'][i])

         entries['TEN HANG HOA'].delete(0,END)
         entries['TEN HANG HOA'].insert(0, str(df['TenHangHoa'][i]) )

         entries['MA VACH'].delete(0,END)
         entries['MA VACH'].insert(0, str(df['MaVach'][i]) )

         entries['GIA CA'].delete(0,END)
         entries['GIA CA'].insert(0, str(df['GiaCa'][i]) )

         return True
         # entries['SO LUONG'].delete(0,END)
         # entries['SO LUONG'].insert(0, str(1) )
         
   return False

init()
ents = makeform(master, fields)
barcode_focus(ents)

b0 = Button(master, text='Cập Nhật Sản Phẩm',font = "Verdana 10 bold",  command=(lambda e=ents: database_modify(e)))
b0.pack(side=TOP)

labels = init_list_detail()

b1 = Button(master, text='Thoát Chương Trình, ESC',font = "Verdana 10 bold", command=onclickQuit)
b1.pack(side=LEFT)

b2 = Button(master, text='Tính Tiền, Enter',font = "Verdana 10 bold", command=(lambda e=ents: sum(e)))
b2.pack(side=RIGHT)

b3 = Button(master, text='Tính Cho Khách Mới, ctrl-Phải',font = "Verdana 10 bold", command=(lambda a=ents, b=labels: reset(a,b)))
b3.pack(side=RIGHT)


t1 = Label(master, 
		 text="TỔNG CỘNG",
		 fg = "blue",
		 bg = "yellow",
		 font = "Verdana 30 bold")
t1.pack()

T = Text(master, height=10, width=80)
T.tag_configure('big', font=('Verdana', 20, 'bold'), justify='center')
T.pack()


master.bind('<Escape>', onclickQuit)
master.bind('<Return>', (lambda event, a=ents, b=labels: enter_pressed(a,b)))
master.bind('<Up>', (lambda event, b=labels : arrow_key_up(b)))
master.bind('<Down>', (lambda event, b=labels : arrow_key_down(b)))
master.bind('<Control_L>', (lambda event, a=ents: database_modify(a)))
master.bind('<Control_R>', (lambda event, a=ents, b=labels : reset(a,b)))



mainloop( )